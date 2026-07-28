from fastapi import FastAPI
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from pydantic import BaseModel, Field, field_validator
from datetime import datetime
from zoneinfo import ZoneInfo
import openpyxl
import os
import re


# =====================================================
# APP
# =====================================================

app = FastAPI(
    title="ClientFlow API",
    version="1.0.0"
)


# =====================================================
# CORS
# =====================================================

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)


# =====================================================
# EXCEL FILE PATH
# =====================================================

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = os.path.join(BASE_DIR, "clients.xlsx")


# =====================================================
# CLIENT DATA MODEL
# =====================================================

class Client(BaseModel):

    portal_id: str = Field(min_length=1, max_length=50)
    first_name: str = Field(min_length=1, max_length=50)
    last_name: str = Field(min_length=1, max_length=50)
    address: str = Field(min_length=1, max_length=100)
    street: str = Field(min_length=1, max_length=100)
    region: str = Field(min_length=1, max_length=50)
    postal_code: str = Field(min_length=6, max_length=6)
    phone_number: str = Field(min_length=10, max_length=10)

    @field_validator("portal_id")
    @classmethod
    def validate_portal_id(cls, value):
        value = value.strip()
        if not re.fullmatch(r"[A-Za-z0-9_-]+", value):
            raise ValueError(
                "Portal ID can contain only letters, numbers, underscore and hyphen"
            )
        return value

    @field_validator("first_name", "last_name", "address", "region")
    @classmethod
    def validate_text_fields(cls, value):
        value = value.strip()
        if not re.fullmatch(r"[A-Za-z ]+", value):
            raise ValueError("This field can contain only alphabets and spaces")
        return value

    @field_validator("street")
    @classmethod
    def validate_street(cls, value):
        value = value.strip()
        if not re.fullmatch(r"[A-Za-z0-9\s\/.,'-]+", value):
            raise ValueError(
                "Street can contain alphabets, numbers, spaces, slash, hyphen, dot, comma and apostrophe"
            )
        return value

    @field_validator("postal_code")
    @classmethod
    def validate_postal_code(cls, value):
        value = value.strip()
        if not re.fullmatch(r"[0-9]{6}", value):
            raise ValueError("Postal code must contain exactly 6 numbers")
        return value

    @field_validator("phone_number")
    @classmethod
    def validate_phone(cls, value):
        value = value.strip()
        if not re.fullmatch(r"[0-9]{10}", value):
            raise ValueError("Phone number must contain exactly 10 digits")
        return value


# =====================================================
# EXCEL HEADER
# =====================================================

EXPECTED_HEADERS = [
    "Portal ID",
    "First Name",
    "Last Name",
    "Address",
    "Street",
    "Region",
    "Postal Code",
    "Phone Number",
    "Created At"
]


# =====================================================
# CREATE EXCEL FILE IF NOT EXISTS
# =====================================================

def create_excel_if_not_exists():

    if not os.path.exists(EXCEL_FILE):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Clients"
        sheet.append(EXPECTED_HEADERS)
        workbook.save(EXCEL_FILE)
        workbook.close()
        return

    # -----------------------------------------------
    # UPDATE OLD EXCEL FILE
    # -----------------------------------------------

    workbook = None

    try:
        workbook = openpyxl.load_workbook(EXCEL_FILE)

        if "Clients" not in workbook.sheetnames:
            sheet = workbook.create_sheet("Clients")
            sheet.append(EXPECTED_HEADERS)
        else:
            sheet = workbook["Clients"]

        existing_headers = [
            sheet.cell(row=1, column=column).value
            for column in range(1, sheet.max_column + 1)
        ]

        # Add missing Postal Code column
        if "Postal Code" not in existing_headers:
            if "Phone Number" in existing_headers:
                phone_index = existing_headers.index("Phone Number") + 1
                sheet.insert_cols(phone_index)
                sheet.cell(row=1, column=phone_index).value = "Postal Code"
            else:
                sheet.cell(row=1, column=sheet.max_column + 1).value = "Postal Code"

        # Add missing Created At column
        existing_headers = [
            sheet.cell(row=1, column=column).value
            for column in range(1, sheet.max_column + 1)
        ]

        if "Created At" not in existing_headers:
            sheet.cell(row=1, column=sheet.max_column + 1).value = "Created At"

        workbook.save(EXCEL_FILE)

    finally:
        if workbook:
            workbook.close()


# =====================================================
# HEALTH CHECK
# =====================================================

@app.get("/")
@app.get("/api")
@app.get("/api/")
@app.get("/health")
def health_check():
    return {
        "success": True,
        "message": "Client Data Automation API is running successfully",
        "timezone": "Asia/Kolkata"
    }


# =====================================================
# ADD CLIENT
# =====================================================

@app.post("/add-client")
def add_client(client: Client):

    try:
        create_excel_if_not_exists()

        workbook = openpyxl.load_workbook(EXCEL_FILE)
        sheet = workbook["Clients"]

        # -------------------------------------------
        # DUPLICATE PORTAL ID CHECK
        # -------------------------------------------

        new_portal_id = client.portal_id.strip().lower()

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue

            existing_portal_id = str(row[0]).strip().lower()

            if existing_portal_id == new_portal_id:
                workbook.close()
                return {
                    "success": False,
                    "message": "Portal ID already exists"
                }

        # -------------------------------------------
        # INDIA TIME
        # -------------------------------------------

        india_time = datetime.now(ZoneInfo("Asia/Kolkata"))
        created_at = india_time.strftime("%d-%m-%Y %I:%M:%S %p")

        # -------------------------------------------
        # SAVE CLIENT
        # -------------------------------------------

        sheet.append([
            client.portal_id,
            client.first_name,
            client.last_name,
            client.address,
            client.street,
            client.region,
            client.postal_code,
            client.phone_number,
            created_at
        ])

        workbook.save(EXCEL_FILE)
        workbook.close()

        return {
            "success": True,
            "message": "Client details saved successfully",
            "created_at": created_at
        }

    except PermissionError:
        return {
            "success": False,
            "message": "clients.xlsx is currently open or locked. Please close Excel and try again."
        }

    except Exception as error:
        return {
            "success": False,
            "message": str(error)
        }


# =====================================================
# GET ALL CLIENTS
# =====================================================

@app.get("/clients")
def get_clients():

    try:
        create_excel_if_not_exists()

        workbook = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
        sheet = workbook["Clients"]

        # -------------------------------------------
        # READ HEADERS
        # -------------------------------------------

        headers = [
            sheet.cell(row=1, column=column).value
            for column in range(1, sheet.max_column + 1)
        ]

        header_index = {
            header: index for index, header in enumerate(headers)
        }

        clients = []

        # -------------------------------------------
        # READ ROWS
        # -------------------------------------------

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue

            def get_value(column_name):
                index = header_index.get(column_name)

                if index is None:
                    return ""

                if index >= len(row):
                    return ""

                value = row[index]

                if value is None:
                    return ""

                return str(value)

            clients.append({
                "portal_id": get_value("Portal ID"),
                "first_name": get_value("First Name"),
                "last_name": get_value("Last Name"),
                "address": get_value("Address"),
                "street": get_value("Street"),
                "region": get_value("Region"),
                "postal_code": get_value("Postal Code"),
                "phone_number": get_value("Phone Number"),
                "created_at": get_value("Created At")
            })

        workbook.close()

        return {
            "success": True,
            "clients": clients,
            "total": len(clients)
        }

    except PermissionError:
        return {
            "success": False,
            "message": "clients.xlsx is currently open or locked. Please close Excel and try again."
        }

    except Exception as error:
        return {
            "success": False,
            "message": str(error)
        }


# =====================================================
# DOWNLOAD EXCEL
# =====================================================

@app.get("/download-excel")
def download_excel():

    create_excel_if_not_exists()

    return FileResponse(
        path=EXCEL_FILE,
        filename="clients.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )